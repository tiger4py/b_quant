# -*- coding: utf-8 -*-
"""
策略交易深度分析工具 — 生成多维度 Excel 分析报告

分析维度：
  1. 概览：核心指标一览
  2. 交易明细：每笔买入/卖出原因 + 卖出后 1-6 月走势
  3. 卖出分析：按卖出类型拆解（止损/移动止盈/到期）
  4. 年度分析：逐年表现
  5. 前向分析：卖出后走势（按卖出类型 + 按月）
  6. 时机分析：卖早了的/卖对了的
  7. 数学分析：分布统计、偏度、峰度、盈亏比分解

用法:
  python script/analyze_trades.py --strategy market_bottom
  python script/analyze_trades.py --strategy market_bottom --json path/to/file.json
"""

import json
import sys
import os
import re
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict, Counter
from statistics import mean, median, stdev, pstdev

ROOT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT_DIR))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from models.stock import StockDaily
from config import DATABASE_URL

# ============ 样式定义 ============

HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FILL_GREEN = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
HEADER_FILL_RED = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
HEADER_FILL_ORANGE = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
CELL_FONT = Font(name='微软雅黑', size=10)
TITLE_FONT = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
SECTION_FONT = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
PROFIT_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
LOSS_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
WARN_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header_row(ws, row, col_start, col_end, fill=HEADER_FILL):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, value=None, fmt=None, font=None, fill=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if fmt:
        cell.number_format = fmt
    cell.font = font or CELL_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    return cell


def get_latest_json(strategy_id):
    """找到最新的回测 JSON 文件"""
    root = ROOT_DIR / 'data/strategy' / strategy_id
    if not root.exists():
        print(f"目录不存在: {root}")
        return None
    month_dirs = sorted([d for d in root.iterdir() if d.is_dir() and re.match(r'^\d{4}-\d{2}$', d.name)],
                        key=lambda d: d.name, reverse=True)
    if not month_dirs:
        return None
    files = sorted(month_dirs[0].glob("*.json"),
                   key=lambda f: f.stat().st_mtime, reverse=True)
    return files[0] if files else None


def classify_sell_reason(reason):
    """分类卖出原因"""
    if not reason:
        return '其他'
    if '期末持仓' in reason:
        return '期末持仓'
    if '止损' in reason:
        return '止损'
    if '移动止盈' in reason or '止盈' in reason:
        return '移动止盈'
    if '到期' in reason:
        return '到期'
    return '其他'


def parse_buy_reason(reason):
    """从买入原因提取关键参数"""
    info = {}
    if not reason:
        return info

    # 提取广度
    m = re.search(r'广度(\d+)%', reason)
    if m: info['breadth'] = int(m.group(1))

    # 提取低于MA百分比
    m = re.search(r'MA\d+(-?[\d.]+)%', reason)
    if m: info['below_ma'] = float(m.group(1))

    # 提取5日跌幅
    m = re.search(r'5日(-?[\d.]+)%', reason)
    if m: info['chg_5d'] = float(m.group(1))

    # 提取ATR
    m = re.search(r'ATR([\d.]+)%', reason)
    if m: info['atr_pct'] = float(m.group(1))

    # 提取止损百分比
    m = re.search(r'止损(-?[\d.]+)%', reason)
    if m: info['stop_pct'] = float(m.group(1))

    return info


def parse_sell_reason(reason):
    """从卖出原因提取关键参数"""
    info = {}
    if not reason:
        return info

    # 提取盈亏百分比
    m = re.search(r'盈(-?[\d.]+)%', reason)
    if m: info['profit_pct'] = float(m.group(1))
    m = re.search(r'\((-?[\d.]+)%', reason)
    if m: info['profit_pct_val'] = float(m.group(1))

    # 提取止损限制
    m = re.search(r'限(-?[\d.]+)%', reason)
    if m: info['stop_limit'] = float(m.group(1))

    # 提取持有天数
    m = re.search(r'持(\d+)天', reason)
    if m: info['hold_days'] = int(m.group(1))

    # 提取高点/回落
    m = re.search(r'高([\d.]+)', reason)
    if m: info['peak'] = float(m.group(1))
    m = re.search(r'回(-?[\d.]+)%', reason)
    if m: info['drawdown'] = float(m.group(1))

    return info


def get_forward_returns(sess, code, sell_date_str, months=[1, 2, 3, 4, 5, 6]):
    """获取卖出后 N 个月的收盘价涨跌"""
    sell_date = datetime.strptime(sell_date_str, '%Y-%m-%d')
    results = {}

    for m in months:
        # 计算目标日期（加 m 个月）
        target_m = sell_date.month + m
        target_y = sell_date.year + (target_m - 1) // 12
        target_m = (target_m - 1) % 12 + 1
        target_d = min(sell_date.day, 28)
        target_date = datetime(target_y, target_m, target_d)

        # 查询 >= 目标日期的第一条记录
        row = sess.query(StockDaily).filter(
            StockDaily.code == code,
            StockDaily.trade_date >= target_date.strftime('%Y-%m-%d')
        ).order_by(StockDaily.trade_date).first()

        if row:
            results[f'{m}m'] = {
                'date': row.trade_date,
                'close': row.close,
            }
        else:
            results[f'{m}m'] = None

    return results


# ============ 主分析函数 ============

def analyze(strategy_id, json_path=None):
    """主入口：加载数据 → 分析 → 生成 Excel"""

    # ----- 1. 加载数据 -----
    if json_path:
        filepath = Path(json_path)
    else:
        filepath = get_latest_json(strategy_id)

    if not filepath or not filepath.exists():
        print(f"找不到回测数据: {filepath}")
        return

    print(f"加载: {filepath}")
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    summary = data['summary']
    trades = data.get('trades', [])

    # 分类交易
    closed_trades = [t for t in trades if classify_sell_reason(t.get('sell_reason', '')) != '期末持仓']
    open_trades = [t for t in trades if classify_sell_reason(t.get('sell_reason', '')) == '期末持仓']

    print(f"总交易: {len(trades)}  已平仓: {len(closed_trades)}  持仓中: {len(open_trades)}")

    # ----- 2. 前向分析（查数据库）-----
    print("查询卖出后走势...")
    engine = create_engine(DATABASE_URL, echo=False)
    Session = sessionmaker(bind=engine)
    sess = Session()

    forward_data = {}
    for t in closed_trades:
        fwd = get_forward_returns(sess, t['code'], t['sell_date'])
        forward_data[t['code'] + '|' + t['buy_date']] = fwd

    # 也查未平仓持仓的买入后走势
    for t in open_trades:
        fwd = get_forward_returns(sess, t['code'], t['buy_date'])
        forward_data[t['code'] + '|' + t['buy_date']] = fwd

    sess.close()

    # ----- 3. 分类统计 -----
    # 按卖出类型
    by_sell_type = defaultdict(list)
    for t in closed_trades:
        cls = classify_sell_reason(t.get('sell_reason', ''))
        by_sell_type[cls].append(t)

    # 按年份
    by_year = defaultdict(list)
    for t in closed_trades:
        by_year[t['buy_date'][:4]].append(t)

    # 按广度分档
    by_breadth = defaultdict(list)
    for t in closed_trades:
        info = parse_buy_reason(t.get('buy_reason', ''))
        breadth = info.get('breadth', 50)
        if breadth < 10:
            bucket = '0-10%'
        elif breadth < 15:
            bucket = '10-15%'
        elif breadth < 20:
            bucket = '15-20%'
        elif breadth < 25:
            bucket = '20-25%'
        else:
            bucket = '>=25%'
        by_breadth[bucket].append(t)

    # ----- 4. 生成 Excel -----
    output_dir = ROOT_DIR / 'data/strategy' / strategy_id / 'analysis'
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{strategy_id}_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = openpyxl.Workbook()

    # ======== Sheet 1: 概览 ========
    ws = wb.active
    ws.title = '概览'

    ws.cell(row=1, column=1, value=f"策略分析报告: {data['strategy']['name']}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"区间: {data['selection']['start_date']} ~ {data['selection']['end_date']}").font = Font(name='微软雅黑', size=10, color='666666')

    # 核心指标
    metrics = [
        ('初始资金', f"{summary['initial_cash']:,.0f}"),
        ('最终权益', f"{summary['final_equity']:,.0f}"),
        ('总收益率', f"{summary['total_return_pct']:.2f}%"),
        ('最大回撤', f"{summary['max_drawdown_pct']:.2f}%"),
        ('交易笔数', f"{summary['trade_count']}"),
        ('已平仓', f"{len(closed_trades)}"),
        ('胜率', f"{summary['win_rate_pct']:.1f}%"),
        ('平均盈利', f"{summary['avg_profit_pct']:.2f}%"),
        ('盈亏因子', f"{summary['profit_factor']:.2f}" if summary.get('profit_factor') else 'N/A'),
        ('最大持仓', f"{summary.get('max_positions', 'N/A')}"),
    ]

    for i, (label, val) in enumerate(metrics):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = Font(name='微软雅黑', bold=True, size=10)
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=val).font = CELL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = CENTER_ALIGN

    # 盈亏分布概览
    r = 4 + len(metrics) + 1
    ws.cell(row=r, column=1, value='盈亏分布').font = SECTION_FONT

    winners = [t for t in closed_trades if t['profit'] > 0]
    losers = [t for t in closed_trades if t['profit'] <= 0]

    dist_rows = [
        ('盈利笔数', len(winners), f"{len(winners)/len(closed_trades)*100:.0f}%"),
        ('亏损笔数', len(losers), f"{len(losers)/len(closed_trades)*100:.0f}%"),
        ('平均盈利(盈)', f"{mean([t['profit_pct'] for t in winners]):.1f}%" if winners else 'N/A'),
        ('平均亏损(亏)', f"{mean([t['profit_pct'] for t in losers]):.1f}%" if losers else 'N/A'),
        ('最大单笔盈利', f"{max([t['profit_pct'] for t in winners]):.1f}%" if winners else 'N/A'),
        ('最大单笔亏损', f"{min([t['profit_pct'] for t in losers]):.1f}%" if losers else 'N/A'),
        ('盈亏比(平均)', f"{abs(mean([t['profit_pct'] for t in winners])/mean([t['profit_pct'] for t in losers])):.2f}" if winners and losers else 'N/A'),
    ]

    r += 1
    for label, *vals in dist_rows:
        ws.cell(row=r, column=1, value=label).font = Font(name='微软雅黑', bold=True, size=10)
        ws.cell(row=r, column=1).border = THIN_BORDER
        for j, v in enumerate(vals):
            style_data_cell(ws, r, 2 + j, v)
        r += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22

    # ======== Sheet 2: 交易明细 ========
    ws2 = wb.create_sheet('交易明细')

    headers2 = ['买入日', '卖出日', '股票', '代码', '买入价', '卖出价', '盈亏%', '盈亏额',
                '持天', '广度%', '低于MA%', '5日跌%', 'ATR%', '止损%',
                '买入原因', '卖出原因', '卖出类型',
                '1月后%', '2月后%', '3月后%', '4月后%', '5月后%', '6月后%', '最佳后续%']

    for j, h in enumerate(headers2):
        ws2.cell(row=1, column=j+1, value=h)
    style_header_row(ws2, 1, 1, len(headers2))

    all_trades = sorted(closed_trades + open_trades, key=lambda x: x['buy_date'])

    for i, t in enumerate(all_trades):
        r = i + 2
        is_open = classify_sell_reason(t.get('sell_reason', '')) == '期末持仓'

        buy_info = parse_buy_reason(t.get('buy_reason', ''))
        sell_info = parse_sell_reason(t.get('sell_reason', ''))

        # 持有天数
        hold_days = ''
        try:
            bd = datetime.strptime(t['buy_date'], '%Y-%m-%d')
            sd = datetime.strptime(t['sell_date'], '%Y-%m-%d')
            hold_days = (sd - bd).days
        except:
            pass

        sell_type = classify_sell_reason(t.get('sell_reason', ''))

        row_data = [
            t['buy_date'], t['sell_date'], t['name'], t['code'],
            t['buy_price'], t['sell_price'],
            t['profit_pct'] / 100,  # 存为数值
            t['profit'],
            hold_days,
            buy_info.get('breadth', ''),
            buy_info.get('below_ma', ''),
            buy_info.get('chg_5d', ''),
            buy_info.get('atr_pct', ''),
            buy_info.get('stop_pct', ''),
            t.get('buy_reason', ''),
            t.get('sell_reason', ''),
            sell_type,
        ]

        for j, v in enumerate(row_data):
            style_data_cell(ws2, r, j + 1, v)

        # 前向收益
        key = t['code'] + '|' + t['buy_date']
        fwd = forward_data.get(key, {})
        best_fwd = -999

        for m_idx, m_label in enumerate(['1m', '2m', '3m', '4m', '5m', '6m']):
            f = fwd.get(m_label)
            if f and t['sell_price'] > 0:
                chg = (f['close'] / t['sell_price'] - 1) * 100
                style_data_cell(ws2, r, 18 + m_idx, f"{chg:+.1f}%",
                               fill=PROFIT_FILL if chg > 0 else LOSS_FILL if chg < 0 else None)
                if chg > best_fwd:
                    best_fwd = chg
            else:
                style_data_cell(ws2, r, 18 + m_idx, '-')

        # 最佳后续
        style_data_cell(ws2, r, 24, f"{best_fwd:+.1f}%" if best_fwd > -999 else '-',
                       fill=PROFIT_FILL if best_fwd > 0 else None)

        # 整行底色
        if not is_open:
            fill = PROFIT_FILL if t['profit'] > 0 else LOSS_FILL
            for j in range(1, 25):
                if ws2.cell(row=r, column=j).fill == PatternFill():  # 没被前向列覆盖的
                    pass  # 前向列已有颜色

        # 盈亏% 格式
        style_data_cell(ws2, r, 7, None, fmt='0.00%')

    # 冻结首行
    ws2.freeze_panes = 'A2'
    # 设置列宽
    for j in range(1, len(headers2) + 1):
        if j in [15, 16]:
            ws2.column_dimensions[get_column_letter(j)].width = 50
        elif j in [18, 19, 20, 21, 22, 23, 24]:
            ws2.column_dimensions[get_column_letter(j)].width = 11
        else:
            ws2.column_dimensions[get_column_letter(j)].width = 14

    # ======== Sheet 3: 卖出分析 ========
    ws3 = wb.create_sheet('卖出分析')

    sell_order = ['止损', '移动止盈', '到期']
    headers3 = ['卖出类型', '笔数', '占比', '胜率', '累计盈亏', '平均盈亏%', '中位数%', '最佳%', '最差%', '平均持天']

    for j, h in enumerate(headers3):
        ws3.cell(row=1, column=j+1, value=h)
    style_header_row(ws3, 1, 1, len(headers3))

    r = 2
    for sell_type in sell_order:
        items = by_sell_type.get(sell_type, [])
        if not items:
            continue
        wins = [t for t in items if t['profit'] > 0]
        pcts = [t['profit_pct'] for t in items]
        hold_days_list = []
        for t in items:
            try:
                bd = datetime.strptime(t['buy_date'], '%Y-%m-%d')
                sd = datetime.strptime(t['sell_date'], '%Y-%m-%d')
                hold_days_list.append((sd - bd).days)
            except:
                pass

        row_data = [
            sell_type,
            len(items),
            len(items) / len(closed_trades),
            len(wins) / len(items),
            sum(t['profit'] for t in items),
            mean(pcts),
            median(pcts),
            max(pcts),
            min(pcts),
            mean(hold_days_list) if hold_days_list else '',
        ]
        for j, v in enumerate(row_data):
            style_data_cell(ws3, r, j + 1, v)
        # 格式化
        style_data_cell(ws3, r, 3, None, fmt='0.0%')
        style_data_cell(ws3, r, 4, None, fmt='0.0%')
        style_data_cell(ws3, r, 6, None, fmt='0.00%')
        style_data_cell(ws3, r, 7, None, fmt='0.00%')
        r += 1

    # 汇总行
    total_wins = [t for t in closed_trades if t['profit'] > 0]
    all_pcts = [t['profit_pct'] for t in closed_trades]
    all_hold = []
    for t in closed_trades:
        try:
            bd = datetime.strptime(t['buy_date'], '%Y-%m-%d')
            sd = datetime.strptime(t['sell_date'], '%Y-%m-%d')
            all_hold.append((sd - bd).days)
        except: pass

    row_data = ['合计', len(closed_trades), 1.0, len(total_wins)/len(closed_trades),
                sum(t['profit'] for t in closed_trades), mean(all_pcts), median(all_pcts),
                max(all_pcts), min(all_pcts), mean(all_hold) if all_hold else '']
    for j, v in enumerate(row_data):
        style_data_cell(ws3, r, j + 1, v, font=Font(name='微软雅黑', bold=True, size=10))

    ws3.column_dimensions['A'].width = 14
    for j in range(2, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(j)].width = 14

    # 止损组详细分析
    r += 2
    ws3.cell(row=r, column=1, value='止损组 - 逐笔详情').font = SECTION_FONT
    r += 1

    stop_headers = ['股票', '买入日', '卖出日', '盈亏%', '广度%', '低于MA%', '买入ATR%', '止损限%',
                    '持天', '1月后%', '2月后%', '3月后%', '6月后%', '最佳后续%', '卖出原因']
    for j, h in enumerate(stop_headers):
        ws3.cell(row=r, column=j+1, value=h)
    style_header_row(ws3, r, 1, len(stop_headers), HEADER_FILL_RED)
    r += 1

    stop_trades = sorted(by_sell_type.get('止损', []), key=lambda x: x['profit_pct'])
    for t in stop_trades:
        buy_info = parse_buy_reason(t.get('buy_reason', ''))
        sell_info = parse_sell_reason(t.get('sell_reason', ''))

        try:
            bd = datetime.strptime(t['buy_date'], '%Y-%m-%d')
            sd = datetime.strptime(t['sell_date'], '%Y-%m-%d')
            hd = (sd - bd).days
        except:
            hd = ''

        key = t['code'] + '|' + t['buy_date']
        fwd = forward_data.get(key, {})

        row_data = [
            t['name'], t['buy_date'], t['sell_date'], t['profit_pct'],
            buy_info.get('breadth', ''),
            buy_info.get('below_ma', ''),
            buy_info.get('atr_pct', ''),
            buy_info.get('stop_pct', ''),
            hd,
        ]
        for j, v in enumerate(row_data):
            style_data_cell(ws3, r, j + 1, v)

        best_fwd = -999
        for m_idx, m_label in enumerate(['1m', '2m', '3m', '6m']):
            f = fwd.get(m_label)
            if f and t['sell_price'] > 0:
                chg = (f['close'] / t['sell_price'] - 1) * 100
                style_data_cell(ws3, r, 10 + m_idx, f"{chg:+.1f}%",
                               fill=PROFIT_FILL if chg > 0 else LOSS_FILL if chg < 0 else None)
                if chg > best_fwd:
                    best_fwd = chg
            else:
                style_data_cell(ws3, r, 10 + m_idx, '-')

        style_data_cell(ws3, r, 14, f"{best_fwd:+.1f}%" if best_fwd > -999 else '-',
                       fill=PROFIT_FILL if best_fwd > 0 else None)
        style_data_cell(ws3, r, 15, t.get('sell_reason', ''))
        r += 1

    for j in range(1, len(stop_headers) + 1):
        ws3.column_dimensions[get_column_letter(j)].width = 13

    # ======== Sheet 4: 年度分析 ========
    ws4 = wb.create_sheet('年度分析')

    headers4 = ['年份', '笔数', '已平仓', '胜率', '累计盈亏', '累计盈亏%', '平均盈亏%', '最佳%', '最差%', '涉及股票数']
    for j, h in enumerate(headers4):
        ws4.cell(row=1, column=j+1, value=h)
    style_header_row(ws4, 1, 1, len(headers4))

    r = 2
    for year in sorted(by_year):
        items = by_year[year]
        closed = [t for t in items if classify_sell_reason(t.get('sell_reason', '')) != '期末持仓']
        if not closed:
            continue
        wins = [t for t in closed if t['profit'] > 0]
        stocks = set(t['code'] for t in closed)

        row_data = [
            year, len(items), len(closed),
            len(wins) / len(closed),
            sum(t['profit'] for t in closed),
            sum(t['profit_pct'] for t in closed),
            mean([t['profit_pct'] for t in closed]),
            max([t['profit_pct'] for t in closed]),
            min([t['profit_pct'] for t in closed]),
            len(stocks),
        ]
        for j, v in enumerate(row_data):
            style_data_cell(ws4, r, j + 1, v)
        style_data_cell(ws4, r, 4, None, fmt='0.0%')
        style_data_cell(ws4, r, 7, None, fmt='0.00%')
        r += 1

    # 逐年前向分析
    r += 1
    ws4.cell(row=r, column=1, value='卖出后走势(按月)').font = SECTION_FONT
    r += 1

    # 按年统计卖出后的平均走势
    year_fwd = defaultdict(lambda: defaultdict(list))
    for t in closed_trades:
        key = t['code'] + '|' + t['buy_date']
        fwd = forward_data.get(key, {})
        year = t['buy_date'][:4]
        for m_label in ['1m', '2m', '3m', '4m', '5m', '6m']:
            f = fwd.get(m_label)
            if f and t['sell_price'] > 0:
                chg = (f['close'] / t['sell_price'] - 1) * 100
                year_fwd[year][m_label].append(chg)

    fwd_headers = ['年份', '笔数'] + [f'{m}月后' for m in range(1, 7)] + ['上涨比']
    for j, h in enumerate(fwd_headers):
        ws4.cell(row=r, column=j+1, value=h)
    style_header_row(ws4, r, 1, len(fwd_headers), HEADER_FILL_GREEN)
    r += 1

    for year in sorted(year_fwd):
        y = year_fwd[year]
        all_vals = []
        row_data = [year, len(by_year.get(year, []))]
        for m_label in ['1m', '2m', '3m', '4m', '5m', '6m']:
            vals = y.get(m_label, [])
            row_data.append(mean(vals) if vals else 0)
            all_vals.extend(vals)
        pos_ratio = sum(1 for v in all_vals if v > 0) / len(all_vals) if all_vals else 0
        row_data.append(pos_ratio)

        for j, v in enumerate(row_data):
            style_data_cell(ws4, r, j + 1, v)
        for j in range(3, 9):
            style_data_cell(ws4, r, j, None, fmt='+0.0%;-0.0%')
        style_data_cell(ws4, r, 9, None, fmt='0%')
        r += 1

    for j in range(1, len(fwd_headers) + 1):
        ws4.column_dimensions[get_column_letter(j)].width = 13

    # ======== Sheet 5: 前向分析（割在地板上）=======
    ws5 = wb.create_sheet('卖早分析')

    ws5.cell(row=1, column=1, value='卖早了 — 卖出后继续大涨的交易').font = TITLE_FONT
    ws5.cell(row=2, column=1, value='判断标准: 卖出后6个月内最高涨幅 > 卖出时盈亏 + 15%').font = Font(name='微软雅黑', size=9, color='666666')

    early_headers = ['股票', '买入日', '卖出日', '卖出时盈亏%', '卖出类型', '广度%',
                     '1月后%', '2月后%', '3月后%', '4月后%', '5月后%', '6月后%',
                     '最佳后续%', '差值(错失pp)', '卖出原因']

    for j, h in enumerate(early_headers):
        ws5.cell(row=4, column=j+1, value=h)
    style_header_row(ws5, 4, 1, len(early_headers), HEADER_FILL_RED)

    # 找出卖早了的
    early_exits = []
    for t in closed_trades:
        key = t['code'] + '|' + t['buy_date']
        fwd = forward_data.get(key, {})
        best_fwd = -999
        for m_label in ['1m', '2m', '3m', '4m', '5m', '6m']:
            f = fwd.get(m_label)
            if f and t['sell_price'] > 0:
                chg = (f['close'] / t['sell_price'] - 1) * 100
                if chg > best_fwd:
                    best_fwd = chg

        if best_fwd > -999 and best_fwd > t['profit_pct'] + 15:
            early_exits.append((t, best_fwd, best_fwd - t['profit_pct']))

    early_exits.sort(key=lambda x: -x[2])  # 按错失幅度排序

    r = 5
    for t, best_fwd, missed in early_exits:
        buy_info = parse_buy_reason(t.get('buy_reason', ''))
        sell_type = classify_sell_reason(t.get('sell_reason', ''))
        key = t['code'] + '|' + t['buy_date']
        fwd = forward_data.get(key, {})

        row_data = [
            t['name'], t['buy_date'], t['sell_date'], t['profit_pct'],
            sell_type,
            buy_info.get('breadth', ''),
        ]
        for j, v in enumerate(row_data):
            style_data_cell(ws5, r, j + 1, v)

        for m_idx, m_label in enumerate(['1m', '2m', '3m', '4m', '5m', '6m']):
            f = fwd.get(m_label)
            if f and t['sell_price'] > 0:
                chg = (f['close'] / t['sell_price'] - 1) * 100
                style_data_cell(ws5, r, 7 + m_idx, f"{chg:+.1f}%",
                               fill=PROFIT_FILL if chg > 0 else LOSS_FILL if chg < 0 else None)
            else:
                style_data_cell(ws5, r, 7 + m_idx, '-')

        style_data_cell(ws5, r, 13, f"{best_fwd:+.1f}%", fill=PROFIT_FILL)
        style_data_cell(ws5, r, 14, f"+{missed:.0f}pp", fill=WARN_FILL)
        style_data_cell(ws5, r, 15, t.get('sell_reason', ''))
        r += 1

    # 统计
    r += 1
    ws5.cell(row=r, column=1, value=f"卖早了: {len(early_exits)}/{len(closed_trades)} = {len(early_exits)/len(closed_trades)*100:.0f}%").font = Font(name='微软雅黑', bold=True, size=11, color='C00000')
    r += 1

    # 按卖出类型统计卖早比例
    for sell_type in ['止损', '移动止盈', '到期']:
        total_in_type = len(by_sell_type.get(sell_type, []))
        early_in_type = len([e for e in early_exits if classify_sell_reason(e[0].get('sell_reason', '')) == sell_type])
        if total_in_type > 0:
            ws5.cell(row=r, column=1, value=f"  {sell_type}: {early_in_type}/{total_in_type} = {early_in_type/total_in_type*100:.0f}%").font = CELL_FONT
            r += 1

    # 最惨TOP 10
    r += 1
    ws5.cell(row=r, column=1, value='割在地板上 TOP 10（错失最大）').font = SECTION_FONT
    r += 1
    for j, h in enumerate(['排名', '股票', '买入日', '卖出日', '卖出时%', '最佳后续%', '错失pp', '卖出类型']):
        ws5.cell(row=r, column=j+1, value=h).font = Font(name='微软雅黑', bold=True, size=9)
    r += 1

    for rank, (t, best_fwd, missed) in enumerate(early_exits[:10], 1):
        sell_type = classify_sell_reason(t.get('sell_reason', ''))
        row_data = [rank, t['name'], t['buy_date'], t['sell_date'], t['profit_pct'], best_fwd, missed, sell_type]
        for j, v in enumerate(row_data):
            style_data_cell(ws5, r, j + 1, v, fill=WARN_FILL if j == 6 else None)
        r += 1

    for j in range(1, len(early_headers) + 1):
        ws5.column_dimensions[get_column_letter(j)].width = 13

    # ======== Sheet 6: 数学分析 ========
    ws6 = wb.create_sheet('数学分析')

    ws6.cell(row=1, column=1, value='统计与分布分析').font = TITLE_FONT

    # 6.1 盈亏分布统计
    r = 3
    ws6.cell(row=r, column=1, value='盈亏分布统计').font = SECTION_FONT
    r += 1

    pcts = [t['profit_pct'] for t in closed_trades]

    stats_rows = [
        ('样本数', len(pcts)),
        ('均值', f"{mean(pcts):.2f}%"),
        ('中位数', f"{median(pcts):.2f}%"),
        ('标准差(样本)', f"{stdev(pcts):.2f}%" if len(pcts) > 1 else 'N/A'),
        ('标准差(总体)', f"{pstdev(pcts):.2f}%" if len(pcts) > 1 else 'N/A'),
        ('最小值', f"{min(pcts):.2f}%"),
        ('最大值', f"{max(pcts):.2f}%"),
        ('极差', f"{max(pcts) - min(pcts):.2f}%"),
        ('偏度', f"{_skewness(pcts):.3f}" if len(pcts) > 2 else 'N/A'),
        ('峰度', f"{_kurtosis(pcts):.3f}" if len(pcts) > 3 else 'N/A'),
        ('夏普近似(均值/标准差)', f"{mean(pcts)/stdev(pcts):.3f}" if len(pcts) > 1 and stdev(pcts) > 0 else 'N/A'),
    ]

    for label, val in stats_rows:
        ws6.cell(row=r, column=1, value=label).font = Font(name='微软雅黑', bold=True, size=10)
        ws6.cell(row=r, column=1).border = THIN_BORDER
        ws6.cell(row=r, column=2, value=val).font = CELL_FONT
        ws6.cell(row=r, column=2).border = THIN_BORDER
        r += 1

    # 6.2 盈亏分桶
    r += 1
    ws6.cell(row=r, column=1, value='盈亏分桶').font = SECTION_FONT
    r += 1

    buckets_def = [
        ('<= -25%', None, -25),
        ('-25% ~ -15%', -25, -15),
        ('-15% ~ -10%', -15, -10),
        ('-10% ~ -5%', -10, -5),
        ('-5% ~ 0%', -5, 0),
        ('0% ~ +5%', 0, 5),
        ('+5% ~ +10%', 5, 10),
        ('+10% ~ +20%', 10, 20),
        ('+20% ~ +50%', 20, 50),
        ('+50% ~ +100%', 50, 100),
        ('> +100%', 100, None),
    ]

    for j, h in enumerate(['区间', '笔数', '占比', '累计盈亏']):
        ws6.cell(row=r, column=j+1, value=h)
    style_header_row(ws6, r, 1, 4)
    r += 1

    for label, lo, hi in buckets_def:
        if lo is None:
            items = [t for t in closed_trades if t['profit_pct'] <= hi]
        elif hi is None:
            items = [t for t in closed_trades if t['profit_pct'] > lo]
        else:
            items = [t for t in closed_trades if lo < t['profit_pct'] <= hi]

        style_data_cell(ws6, r, 1, label)
        style_data_cell(ws6, r, 2, len(items))
        style_data_cell(ws6, r, 3, len(items) / len(closed_trades) if closed_trades else 0, fmt='0.0%')
        style_data_cell(ws6, r, 4, sum(t['profit'] for t in items))

        # 高亮
        if items:
            avg_p = mean([t['profit_pct'] for t in items])
            if avg_p > 0:
                ws6.cell(row=r, column=1).fill = PROFIT_FILL
            elif avg_p < 0:
                ws6.cell(row=r, column=1).fill = LOSS_FILL
        r += 1

    # 6.3 广度 vs 盈亏
    r += 1
    ws6.cell(row=r, column=1, value='广度分档 vs 盈亏').font = SECTION_FONT
    r += 1

    breadth_headers = ['广度区间', '笔数', '胜率', '平均盈亏%', '中位数%', '最佳%', '最差%']
    for j, h in enumerate(breadth_headers):
        ws6.cell(row=r, column=j+1, value=h)
    style_header_row(ws6, r, 1, len(breadth_headers), HEADER_FILL_ORANGE)
    r += 1

    breadth_order = ['0-10%', '10-15%', '15-20%', '20-25%', '>=25%']
    for bucket in breadth_order:
        items = by_breadth.get(bucket, [])
        if not items:
            continue
        wins = [t for t in items if t['profit'] > 0]
        row_data = [
            bucket, len(items),
            len(wins) / len(items),
            mean([t['profit_pct'] for t in items]),
            median([t['profit_pct'] for t in items]),
            max([t['profit_pct'] for t in items]),
            min([t['profit_pct'] for t in items]),
        ]
        for j, v in enumerate(row_data):
            style_data_cell(ws6, r, j + 1, v)
        style_data_cell(ws6, r, 3, None, fmt='0.0%')
        r += 1

    # 6.4 低于MA vs 盈亏
    r += 1
    ws6.cell(row=r, column=1, value='超跌程度 vs 盈亏').font = SECTION_FONT
    r += 1

    ma_buckets = defaultdict(list)
    for t in closed_trades:
        info = parse_buy_reason(t.get('buy_reason', ''))
        bma = info.get('below_ma', 0)
        if bma < -20:
            b = '< -20%'
        elif bma < -15:
            b = '-20% ~ -15%'
        elif bma < -10:
            b = '-15% ~ -10%'
        elif bma < -8:
            b = '-10% ~ -8%'
        else:
            b = '> -8%'
        ma_buckets[b].append(t)

    ma_headers = ['低于MA60', '笔数', '胜率', '平均盈亏%', '中位数%']
    for j, h in enumerate(ma_headers):
        ws6.cell(row=r, column=j+1, value=h)
    style_header_row(ws6, r, 1, len(ma_headers), HEADER_FILL_GREEN)
    r += 1

    ma_order = ['< -20%', '-20% ~ -15%', '-15% ~ -10%', '-10% ~ -8%', '> -8%']
    for bucket in ma_order:
        items = ma_buckets.get(bucket, [])
        if not items:
            continue
        wins = [t for t in items if t['profit'] > 0]
        row_data = [bucket, len(items), len(wins) / len(items),
                    mean([t['profit_pct'] for t in items]),
                    median([t['profit_pct'] for t in items])]
        for j, v in enumerate(row_data):
            style_data_cell(ws6, r, j + 1, v)
        style_data_cell(ws6, r, 3, None, fmt='0.0%')
        r += 1

    # 6.5 Top N 集中度
    r += 1
    ws6.cell(row=r, column=1, value='利润集中度').font = SECTION_FONT
    r += 1

    winners_sorted = sorted([t for t in closed_trades if t['profit'] > 0], key=lambda x: -x['profit'])
    total_winner_profit = sum(t['profit'] for t in winners_sorted)

    if total_winner_profit > 0:
        for top_n in [3, 5, 10]:
            top_profit = sum(t['profit'] for t in winners_sorted[:top_n])
            pct = top_profit / total_winner_profit * 100
            ws6.cell(row=r, column=1, value=f"Top {top_n} 赢家贡献").font = Font(name='微软雅黑', bold=True, size=10)
            ws6.cell(row=r, column=2, value=f"{top_profit:,.0f} / {total_winner_profit:,.0f} = {pct:.0f}%").font = CELL_FONT
            r += 1

    for j in range(1, 5):
        ws6.column_dimensions[get_column_letter(j)].width = 18

    # ======== Sheet 7: 优化建议汇总 ========
    ws7 = wb.create_sheet('优化建议')

    ws7.cell(row=1, column=1, value='策略优化建议（自动生成）').font = TITLE_FONT
    ws7.cell(row=2, column=1, value='基于数据分析的优化方向，按优先级排列').font = Font(name='微软雅黑', size=9, color='666666')

    suggestions = _generate_suggestions(closed_trades, by_sell_type, forward_data, by_year, by_breadth, early_exits)

    r = 4
    for i, (priority, title, detail, evidence) in enumerate(suggestions, 1):
        fill = HEADER_FILL_RED if priority == 'P0' else HEADER_FILL_ORANGE if priority == 'P1' else HEADER_FILL_GREEN

        ws7.cell(row=r, column=1, value=f"[{priority}] {title}").font = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
        ws7.cell(row=r, column=1).fill = fill
        ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

        ws7.cell(row=r, column=1, value='问题描述:').font = Font(name='微软雅黑', bold=True, size=10)
        ws7.cell(row=r, column=2, value=detail).font = CELL_FONT
        ws7.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        r += 1

        ws7.cell(row=r, column=1, value='数据支撑:').font = Font(name='微软雅黑', bold=True, size=10)
        ws7.cell(row=r, column=2, value=evidence).font = CELL_FONT
        ws7.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        r += 2

    ws7.column_dimensions['A'].width = 14
    ws7.column_dimensions['B'].width = 60
    ws7.column_dimensions['C'].width = 20

    # ----- 5. 保存 -----
    wb.save(output_path)
    print(f"\n分析报告已生成: {output_path}")
    return output_path


def _skewness(data):
    """计算偏度"""
    n = len(data)
    if n < 3:
        return 0
    m = mean(data)
    s = stdev(data)
    if s == 0:
        return 0
    return sum(((x - m) / s) ** 3 for x in data) * n / ((n - 1) * (n - 2))


def _kurtosis(data):
    """计算超额峰度"""
    n = len(data)
    if n < 4:
        return 0
    m = mean(data)
    s = stdev(data)
    if s == 0:
        return 0
    k = sum(((x - m) / s) ** 4 for x in data) * n * (n + 1) / ((n - 1) * (n - 2) * (n - 3))
    return k - 3 * (n - 1) ** 2 / ((n - 2) * (n - 3))


def _generate_suggestions(closed_trades, by_sell_type, forward_data, by_year, by_breadth, early_exits):
    """基于数据自动生成优化建议"""
    suggestions = []

    if not closed_trades:
        suggestions.append(('P0', '无交易数据', '策略未产生任何已平仓交易', ''))
        return suggestions

    # ---- 止损分析 ----
    stop_trades = by_sell_type.get('止损', [])
    if stop_trades:
        stop_pcts = [t['profit_pct'] for t in stop_trades]
        avg_stop_loss = mean(stop_pcts)
        stop_ratio = len(stop_trades) / len(closed_trades)

        # 检查止损后走势
        stop_fwd_best = []
        for t in stop_trades:
            key = t['code'] + '|' + t['buy_date']
            fwd = forward_data.get(key, {})
            best = -999
            for m_label in ['1m', '2m', '3m', '6m']:
                f = fwd.get(m_label)
                if f and t['sell_price'] > 0:
                    chg = (f['close'] / t['sell_price'] - 1) * 100
                    if chg > best:
                        best = chg
            if best > -999:
                stop_fwd_best.append((t, best))

        stop_recovered = [(t, b) for t, b in stop_fwd_best if b > 0]
        stop_recovered_big = [(t, b) for t, b in stop_fwd_best if b > 10]

        if stop_recovered:
            suggestions.append((
                'P0',
                f'止损后反弹 — {len(stop_recovered)}/{len(stop_fwd_best)}笔止损后反弹',
                f'止损组 {len(stop_trades)}笔交易中，{len(stop_recovered)}笔({len(stop_recovered)/max(len(stop_fwd_best),1)*100:.0f}%) 卖出后价格回升为正。'
                f'其中 {len(stop_recovered_big)}笔回升超10%，说明止损可能偏紧。',
                f'平均止损亏损: {avg_stop_loss:.1f}% | 止损占比: {stop_ratio*100:.0f}% | '
                f'止损后1月平均: {_avg_fwd_at_month(stop_trades, forward_data, 1):+.1f}% | '
                f'3月平均: {_avg_fwd_at_month(stop_trades, forward_data, 3):+.1f}%'
            ))

        # 跳空穿止损
        big_losers = [t for t in stop_trades if t['profit_pct'] < -18]
        if big_losers:
            suggestions.append((
                'P1',
                f'跳空穿止损 — {len(big_losers)}笔亏损超18%',
                f'部分止损单亏损远超止损线，可能是跳空低开或连续跌停导致。考虑加入"一字跌停无法卖出"的模拟逻辑。',
                f'最大跳空: {min([t["profit_pct"] for t in big_losers]):.1f}%'
            ))

    # ---- 胜率分析 ----
    total_wins = sum(1 for t in closed_trades if t['profit'] > 0)
    win_rate = total_wins / len(closed_trades) if closed_trades else 0
    if win_rate < 0.4:
        suggestions.append((
            'P1',
            f'胜率偏低({win_rate*100:.0f}%)',
            f'当前胜率 {win_rate*100:.0f}% 低于40%，接近一半交易亏损。考虑: (1)收紧选股条件 (2)增加确认信号 (3)用广度分级控制仓位',
            f'盈利交易: {total_wins}笔 | 亏损: {len(closed_trades) - total_wins}笔'
        ))

    # ---- 年度分布 ----
    year_returns = {}
    for year in sorted(by_year):
        closed = [t for t in by_year[year] if t.get('sell_reason', '') != '期末持仓']
        if closed:
            year_returns[year] = sum(t['profit_pct'] for t in closed)

    neg_years = [y for y, r in year_returns.items() if r < 0]
    if len(neg_years) >= 2:
        suggestions.append((
            'P1',
            f'{len(neg_years)}年亏损',
            f'连续多年亏损表明策略在某些市场环境下失效。分析亏损年份的市场特征(牛市/熊市/震荡)，针对性加过滤。',
            f'亏损年份: {", ".join(neg_years)} | 年收益: {", ".join(f"{y}:{year_returns[y]:.0f}%" for y in neg_years)}'
        ))

    # ---- 盈亏比 ----
    winners = [t for t in closed_trades if t['profit'] > 0]
    losers = [t for t in closed_trades if t['profit'] <= 0]
    if winners and losers:
        avg_win = mean([t['profit_pct'] for t in winners])
        avg_loss = abs(mean([t['profit_pct'] for t in losers]))
        rr_ratio = avg_win / avg_loss if avg_loss > 0 else 0
        if rr_ratio < 1.5:
            suggestions.append((
                'P1',
                f'盈亏比偏低({rr_ratio:.2f})',
                f'平均盈利 {avg_win:.1f}% vs 平均亏损 {avg_loss:.1f}%，盈亏比 {rr_ratio:.2f}。'
                f'需要让盈利跑得更远(放宽止盈)或控制亏损(收紧止损)。',
                f'盈利均值: {avg_win:.2f}% | 亏损均值: -{avg_loss:.2f}%'
            ))

    # ---- 广度分析 ----
    for bucket in ['0-10%', '10-15%']:
        items = by_breadth.get(bucket, [])
        if items:
            avg_r = mean([t['profit_pct'] for t in items])
            if avg_r < 0:
                suggestions.append((
                    'P2',
                    f'极度恐慌({bucket})反而亏钱',
                    f'广度 < 15% 的极端恐慌环境下，平均盈亏 {avg_r:.1f}%，'
                    f'说明极端行情下当前策略可能无法盈利。考虑在广度<10%时暂停交易或减少仓位。',
                    f'{bucket}: {len(items)}笔, 平均{avg_r:.1f}%'
                ))

    # ---- Top N 集中度 ----
    if winners:
        top3 = sorted(winners, key=lambda x: -x['profit'])[:3]
        top3_pct = sum(t['profit'] for t in top3) / sum(t['profit'] for t in winners) * 100
        if top3_pct > 50:
            suggestions.append((
                'P2',
                f'利润高度集中 — Top 3 贡献 {top3_pct:.0f}%',
                f'前3大赢家贡献了 {top3_pct:.0f}% 的总利润，策略极度依赖少数大赢家。'
                f'样本量不够时需警惕过拟合。考虑增加交易频率或降低单笔集中度。',
                f'Top 1: {top3[0]["name"]} +{top3[0]["profit_pct"]:.1f}% | '
                f'Top 2: {top3[1]["name"]} +{top3[1]["profit_pct"]:.1f}% | '
                f'Top 3: {top3[2]["name"]} +{top3[2]["profit_pct"]:.1f}%'
            ))

    return suggestions


def _avg_fwd_at_month(trades, forward_data, month):
    """计算某组交易在卖出后第N个月的平均涨跌"""
    m_label = f'{month}m'
    vals = []
    for t in trades:
        key = t['code'] + '|' + t['buy_date']
        fwd = forward_data.get(key, {})
        f = fwd.get(m_label)
        if f and t['sell_price'] > 0:
            vals.append((f['close'] / t['sell_price'] - 1) * 100)
    return mean(vals) if vals else 0


# ============ CLI ============

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='策略交易深度分析')
    parser.add_argument('--strategy', type=str, default='market_bottom', help='策略ID')
    parser.add_argument('--json', type=str, default=None, help='指定JSON文件路径')
    args = parser.parse_args()

    analyze(args.strategy, args.json)
